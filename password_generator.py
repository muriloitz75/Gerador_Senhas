# -*- coding: utf-8 -*-
import os
import random
import string
import sys
from datetime import datetime
from pathlib import Path
import pyperclip
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                            QLabel, QCheckBox, QSpinBox, QLineEdit, QPushButton,
                            QMessageBox, QInputDialog, QGraphicsDropShadowEffect,
                            QGroupBox, QShortcut)
from PyQt5.QtCore import Qt  # Removida a importação do QKeySequence daqui
from PyQt5.QtGui import QIcon, QPixmap, QColor, QFont, QKeySequence
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class PasswordGenerator(QMainWindow):
    def __init__(self):
        super().__init__()
        
        # Inicializa o histórico de senhas
        self.password_history = []
        
        # Configuração básica da janela
        self.setWindowTitle(u"Gerador de Senhas Profissional")
        self.setFixedSize(700, 500)
        
        # Configurar o ícone da janela e imagem
        try:
            # Obtém o diretório do script
            current_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Tenta carregar o ícone da janela
            icon_path = os.path.join(current_dir, 'ico1.ico')
            if os.path.exists(icon_path):
                self.setWindowIcon(QIcon(icon_path))
            
            # Tenta carregar a imagem
            img_path = os.path.join(current_dir, 'img.png')
            if os.path.exists(img_path):
                icon_pixmap = QPixmap(img_path)
                if icon_pixmap.isNull():
                    print(f"Erro ao carregar imagem: {img_path}")
                    icon_pixmap = QPixmap()
            else:
                print(f"Arquivo de imagem não encontrado: {img_path}")
                icon_pixmap = QPixmap()
        
        except Exception as e:
            print(f"Erro ao carregar recursos: {str(e)}")
            icon_pixmap = QPixmap()
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # Container para ícone e título (agora vertical)
        header_container = QVBoxLayout()
        
        # Ícone/Imagem (centralizado)
        icon_label = QLabel()
        if not icon_pixmap.isNull():
            scaled_pixmap = icon_pixmap.scaled(
                64, 64,  # Mantém o tamanho de 64x64
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            icon_label.setPixmap(scaled_pixmap)
        icon_label.setAlignment(Qt.AlignCenter)
        icon_label.setMinimumSize(64, 64)
        icon_label.setMaximumSize(64, 64)
        icon_label.setStyleSheet("""
            QLabel {
                background-color: transparent;
                border: none;
            }
        """)
        
        # Container para centralizar o ícone
        icon_center_container = QHBoxLayout()
        icon_center_container.addStretch()
        icon_center_container.addWidget(icon_label)
        icon_center_container.addStretch()
        header_container.addLayout(icon_center_container)
        
        # Título (centralizado)
        title = QLabel(u"GERADOR DE SENHAS")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("""
            QLabel {
                font-size: 26px;
                color: #3498DB;
                font-weight: bold;
                font-family: 'Segoe UI', Arial;
                padding: 0;
                margin: 0;
            }
        """)
        header_container.addWidget(title)
        
        # Adiciona espaçamento antes e depois do header
        layout.addSpacing(20)
        layout.addLayout(header_container)
        layout.addSpacing(20)
        
        # Opções de caracteres
        options_group = QGroupBox("Opções de Caracteres")
        options_layout = QVBoxLayout()
        
        self.use_letters = QCheckBox("Letras (a-z, A-Z)")
        self.use_numbers = QCheckBox("Números (0-9)")
        self.use_symbols = QCheckBox("Símbolos (!@#$%^&*)")
        
        # Marcando as opções por padrão
        self.use_letters.setChecked(True)
        self.use_numbers.setChecked(True)
        self.use_symbols.setChecked(True)
        
        options_layout.addWidget(self.use_letters)
        options_layout.addWidget(self.use_numbers)
        options_layout.addWidget(self.use_symbols)
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        
        # Comprimento da senha
        length_container = QHBoxLayout()
        length_label = QLabel(u"Comprimento da Senha:")
        self.length_spin = QSpinBox()
        self.length_spin.setRange(4, 64)
        self.length_spin.setValue(12)
        self.length_spin.setStyleSheet("background: #2C3E50; color: white; padding: 5px;")
        length_container.addWidget(length_label)
        length_container.addWidget(self.length_spin)
        length_container.addStretch()
        layout.addLayout(length_container)
        
        # Campo de senha
        password_label = QLabel(u"Senha Gerada:")
        self.password_field = QLineEdit()
        self.password_field.setPlaceholderText(u"Sua senha aparecerá aqui")
        self.password_field.setReadOnly(True)
        layout.addWidget(password_label)
        layout.addWidget(self.password_field)
        
        # Botões
        button_layout = QHBoxLayout()
        self.generate_btn = QPushButton(u"Gerar Senha")
        self.save_excel_btn = QPushButton(u"Salvar na Planilha")
        self.copy_btn = QPushButton(u"Copiar")
        
        # Estilização moderna dos botões
        button_style = """
            QPushButton {
                border-radius: 8px;
                padding: 12px 25px;
                font-size: 14px;
                font-weight: bold;
                font-family: 'Segoe UI', Arial;
                color: white;
                border: none;
            }
            QPushButton:hover {
                opacity: 0.8;
            }
            QPushButton:pressed {
                padding: 12px 25px;
                opacity: 1;
            }
        """
        
        self.generate_btn.setStyleSheet(button_style + """
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                          stop:0 #E74C3C, stop:1 #C0392B);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                          stop:0 #C0392B, stop:1 #E74C3C);
            }
        """)
        
        self.save_excel_btn.setStyleSheet(button_style + """
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                          stop:0 #3498DB, stop:1 #2980B9);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                          stop:0 #2980B9, stop:1 #3498DB);
            }
        """)
        
        self.copy_btn.setStyleSheet(button_style + """
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                          stop:0 #2ECC71, stop:1 #27AE60);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                          stop:0 #27AE60, stop:1 #2ECC71);
            }
        """)
        
        # Adiciona sombra aos botões
        for btn in [self.generate_btn, self.save_excel_btn, self.copy_btn]:
            effect = QGraphicsDropShadowEffect(self)
            effect.setBlurRadius(10)
            effect.setOffset(0, 3)
            effect.setColor(QColor(0, 0, 0, 50))
            btn.setGraphicsEffect(effect)
        
        # Conectando os botões às funções
        self.generate_btn.clicked.connect(self.generate_password)
        self.save_excel_btn.clicked.connect(self.save_to_excel)
        self.copy_btn.clicked.connect(self.copy_password)
        
        # Adiciona espaçamento entre os botões
        button_layout.addStretch()
        button_layout.addWidget(self.generate_btn)
        button_layout.addSpacing(15)
        button_layout.addWidget(self.save_excel_btn)
        button_layout.addSpacing(15)
        button_layout.addWidget(self.copy_btn)
        button_layout.addSpacing(15)
        
        # Botão de limpar registros antigos
        self.clean_old_btn = QPushButton("Clean_Old")
        self.clean_old_btn.setStyleSheet(button_style + """
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                          stop:0 #95A5A6, stop:1 #7F8C8D);
                padding: 8px 15px;  /* Botão um pouco menor */
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                          stop:0 #7F8C8D, stop:1 #95A5A6);
            }
        """)
        
        # Adicione o efeito de sombra também para o novo botão
        effect = QGraphicsDropShadowEffect(self)
        effect.setBlurRadius(10)
        effect.setOffset(0, 3)
        effect.setColor(QColor(0, 0, 0, 50))
        self.clean_old_btn.setGraphicsEffect(effect)
        
        # Conecte o botão à sua função
        self.clean_old_btn.clicked.connect(self.clean_old_records)
        
        # Adicione o botão ao layout (após o botão de copiar e antes do botão de tema)
        button_layout.addWidget(self.clean_old_btn)
        button_layout.addSpacing(15)
        
        # Botão de tema
        self.theme_btn = QPushButton()
        self.dark_mode = False
        self.theme_btn.clicked.connect(self.toggle_theme)
        self.update_theme_button()
        button_layout.addWidget(self.theme_btn)
        
        layout.addLayout(button_layout)
        
        # Adiciona espaço expansível antes dos créditos
        layout.addStretch()
        
        # Créditos no rodapé (alinhado à direita)
        credits = QLabel("Criado por Murilo Miguel ®")
        credits.setAlignment(Qt.AlignRight | Qt.AlignBottom)
        credits.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7F8C8D;
                font-family: 'Segoe UI', Arial;
                padding: 10px;
                margin: 0;
            }
        """)
        layout.addWidget(credits)

        # Atalhos de teclado
        QShortcut(QKeySequence("Ctrl+G"), self, self.generate_password)
        QShortcut(QKeySequence("Ctrl+S"), self, self.save_to_excel)
        QShortcut(QKeySequence("Ctrl+C"), self, self.copy_password)
        QShortcut(QKeySequence("Ctrl+H"), self, self.show_history)

    def generate_password(self):
        """Gera uma nova senha com base nas opções selecionadas"""
        # Verifica se pelo menos uma opção está selecionada
        if not any([self.use_letters.isChecked(), 
                    self.use_numbers.isChecked(), 
                    self.use_symbols.isChecked()]):
            QMessageBox.warning(self, "Aviso", "Selecione pelo menos um tipo de caractere!")
            return

        # Define os caracteres possíveis com base nas opções selecionadas
        chars = ''
        if self.use_letters.isChecked():
            chars += string.ascii_letters
        if self.use_numbers.isChecked():
            chars += string.digits
        if self.use_symbols.isChecked():
            chars += "!@#$%^&*()_+-=[]{}|;:,.<>?"

        try:
            # Gera a senha
            length = self.length_spin.value()
            password = ''.join(random.choice(chars) for _ in range(length))
            
            # Garante que pelo menos um caractere de cada tipo selecionado está presente
            while self.use_letters.isChecked() and not any(c.isalpha() for c in password) or \
                  self.use_numbers.isChecked() and not any(c.isdigit() for c in password) or \
                  self.use_symbols.isChecked() and not any(c in "!@#$%^&*()_+-=[]{}|;:,.<>?" for c in password):
                password = ''.join(random.choice(chars) for _ in range(length))
            
            self.password_field.setText(password)
            
            # Adicionar ao histórico
            if self.password_field.text():
                self.password_history.append({
                    'senha': self.password_field.text(),
                    'data': datetime.now().strftime("%H:%M:%S"),
                    'comprimento': self.length_spin.value()
                })
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar senha: {str(e)}")

    def save_password(self):
        """Salva a senha gerada em um arquivo com detalhes"""
        try:
            if not self.password_field.text():
                QMessageBox.warning(self, "Aviso", "Gere uma senha primeiro!")
                return

            # Diálogo para obter o nome/descrição da senha
            nome, ok = QInputDialog.getText(
                self, 
                'Salvar Senha', 
                'Digite um nome ou descrição para esta senha:',
                QLineEdit.Normal
            )
            
            if not ok or not nome.strip():
                return

            # Obtém a data e hora atual formatada
            data_hora = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
            
            # Prepara o conteúdo a ser salvo
            conteudo = f"""DETALHES DA SENHA
----------------------------------------
Nome/Descrição: {nome}
Data de Geração: {data_hora}
----------------------------------------
SENHA: {self.password_field.text()}
----------------------------------------
CONFIGURAÇÕES UTILIZADAS:
- Comprimento: {self.length_spin.value()} caracteres
- Letras (a-z, A-Z): {'Sim' if self.use_letters.isChecked() else 'Não'}
- Números (0-9): {'Sim' if self.use_numbers.isChecked() else 'Não'}
- Símbolos especiais: {'Sim' if self.use_symbols.isChecked() else 'Não'}
----------------------------------------"""

            # Nome padrão do arquivo baseado no nome fornecido e data
            nome_arquivo = f"{nome}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            # Remove caracteres inválidos do nome do arquivo
            nome_arquivo = "".join(c for c in nome_arquivo if c.isalnum() or c in "._- ")

            # Garante que temos um nome de arquivo válido
            if not nome_arquivo:
                nome_arquivo = f"senha_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Salvar Senha",
                nome_arquivo,
                "Arquivo de Texto (*.txt);;Todos os Arquivos (*.*)"
            )

            if file_name:
                # Garante que o arquivo terá extensão .txt
                if not file_name.lower().endswith('.txt'):
                    file_name += '.txt'
                
                with open(file_name, 'w', encoding='utf-8') as file:
                    file.write(conteudo)
                QMessageBox.information(self, "Sucesso", "Senha salva com sucesso!")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar a senha: {str(e)}")

    def copy_password(self):
        """Copia a senha para a área de transferência"""
        if not self.password_field.text():
            QMessageBox.warning(self, "Aviso", "Gere uma senha primeiro!")
            return
        pyperclip.copy(self.password_field.text())
        # Feedback visual opcional
        self.statusBar().showMessage("Senha copiada!", 2000)  # Mostra por 2 segundos

    def save_to_excel(self):
        """Salva a senha em uma planilha Excel"""
        if not self.password_field.text():
            QMessageBox.warning(self, "Aviso", "Gere uma senha primeiro!")
            return

        nome, ok = QInputDialog.getText(
            self, 
            'Salvar Senha', 
            'Digite um nome ou descrição para esta senha:',
            QLineEdit.Normal
        )
        
        if not ok or not nome.strip():
            return

        try:
            excel_file = Path('senhas_database.xlsx')
            
            # Cria uma nova planilha ou carrega a existente
            if excel_file.exists():
                wb = load_workbook(excel_file)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                # Adiciona cabeçalho
                ws.append(["ID", "Nome/Descrição", "Senha", "Data Criação", "Hora Criação", "Dias", "Configurações"])
                # Formata cabeçalho
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
            
            # Prepara os dados
            data_atual = datetime.now().strftime("%d/%m/%Y")
            hora_atual = datetime.now().strftime("%H:%M:%S")
            configuracoes = (
                f"Comprimento: {self.length_spin.value()}, "
                f"Letras: {'Sim' if self.use_letters.isChecked() else 'Não'}, "
                f"Números: {'Sim' if self.use_numbers.isChecked() else 'Não'}, "
                f"Símbolos: {'Sim' if self.use_symbols.isChecked() else 'Não'}"
            )
            
            # Adiciona nova linha
            proxima_linha = ws.max_row + 1
            nova_linha = [
                ws.max_row,  # ID
                nome.strip(),
                self.password_field.text(),
                data_atual,
                hora_atual,
                f'=DATEDIF(DATE(MID(D{proxima_linha},7,4),MID(D{proxima_linha},4,2),LEFT(D{proxima_linha},2)),TODAY(),"D")',
                configuracoes
            ]
            ws.append(nova_linha)

            # Ajusta largura das colunas
            ws.column_dimensions['A'].width = 5   # ID
            ws.column_dimensions['B'].width = 20  # Nome/Descrição
            ws.column_dimensions['C'].width = 25  # Senha
            ws.column_dimensions['D'].width = 12  # Data Criação
            ws.column_dimensions['E'].width = 10  # Hora Criação
            ws.column_dimensions['F'].width = 8   # Dias
            ws.column_dimensions['G'].width = 60  # Configurações

            # Centraliza algumas colunas
            for row in ws.iter_rows(min_row=2):
                row[0].alignment = Alignment(horizontal='center')  # ID
                row[3].alignment = Alignment(horizontal='center')  # Data
                row[4].alignment = Alignment(horizontal='center')  # Hora
                row[5].alignment = Alignment(horizontal='center')  # Dias

            # Salva a planilha
            wb.save(excel_file)
            QMessageBox.information(self, "Sucesso", "Senha salva com sucesso!")

            # Tenta abrir a planilha
            try:
                if sys.platform == 'win32':
                    os.startfile(excel_file)
                elif sys.platform == 'darwin':
                    os.system(f'open "{excel_file}"')
                else:
                    os.system(f'xdg-open "{excel_file}"')
            except:
                pass

        except PermissionError:
            QMessageBox.critical(self, "Erro", 
                               "Não foi possível salvar. Verifique se a planilha está aberta em outro programa.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar: {str(e)}")

    def create_new_workbook(self):
        """Cria uma nova planilha com o cabeçalho formatado"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Senhas"

        # Define o cabeçalho
        headers = ["ID", "Nome/Descrição", "Senha", "Data/Hora", "Dias", "Configurações"]
        ws.append(headers)

        # Formata o cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Define larguras iniciais das colunas
        ws.column_dimensions['A'].width = 5   # ID
        ws.column_dimensions['B'].width = 20  # Nome/Descrição
        ws.column_dimensions['C'].width = 25  # Senha
        ws.column_dimensions['D'].width = 20  # Data/Hora
        ws.column_dimensions['E'].width = 15  # Dias
        ws.column_dimensions['F'].width = 70  # Configurações

        return wb

    def check_password_strength(self, password):
        """Avalia a força da senha"""
        score = 0
        feedback = []
        
        # Comprimento
        if len(password) >= 12:
            score += 2
        elif len(password) >= 8:
            score += 1
        else:
            feedback.append("Senha muito curta")
        
        # Complexidade
        if any(c.isupper() for c in password): score += 1
        if any(c.islower() for c in password): score += 1
        if any(c.isdigit() for c in password): score += 1
        if any(c in string.punctuation for c in password): score += 1
        
        # Feedback baseado no score
        if score < 2:
            return "Fraca", "red"
        elif score < 3:
            return "Média", "orange"
        elif score < 5:
            return "Forte", "green"
        else:
            return "Muito Forte", "darkgreen"

    def update_password_field(self):
        """Atualiza o campo de senha com indicador de força"""
        password = self.password_field.text()
        if password:
            strength, color = self.check_password_strength(password)
            self.password_field.setStyleSheet(f"color: {color}; font-weight: bold;")
            self.strength_label.setText(f"Força: {strength}")
            self.strength_label.setStyleSheet(f"color: {color}; font-weight: bold;")

    def show_history(self):
        """Mostra histórico de senhas geradas na sessão"""
        if not self.password_history:
            QMessageBox.information(self, "Histórico", "Nenhuma senha foi gerada ainda.")
            return
            
        history_text = "Senhas geradas nesta sessão:\n\n"
        for entry in reversed(self.password_history[-10:]):  # últimas 10 senhas
            history_text += f"[{entry['data']}] {entry['senha']} ({entry['comprimento']} caracteres)\n"
            
        QMessageBox.information(self, "Histórico", history_text)

    def toggle_theme(self):
        """Alterna entre tema claro e escuro"""
        self.dark_mode = not self.dark_mode
        self.apply_theme()
        self.update_theme_button()
        
    def apply_theme(self):
        """Aplica o tema selecionado"""
        if self.dark_mode:
            self.setStyleSheet("""
                QMainWindow, QWidget {
                    background-color: #2C3E50;
                    color: #ECF0F1;
                }
                QGroupBox {
                    border: 2px solid #34495E;
                    border-radius: 5px;
                    margin-top: 1em;
                    padding-top: 10px;
                }
                QLineEdit, QSpinBox {
                    background-color: #34495E;
                    color: #ECF0F1;
                    border: 1px solid #7F8C8D;
                    padding: 5px;
                    border-radius: 3px;
                }
            """)
        else:
            self.setStyleSheet("")  # Reset para tema padrão
            
    def update_theme_button(self):
        """Atualiza o ícone e texto do botão de tema"""
        self.theme_btn.setText("🌙" if self.dark_mode else "☀️")
        self.theme_btn.setToolTip("Mudar para tema claro" if self.dark_mode else "Mudar para tema escuro")

    def clear_password_field(self):
        """Limpa o campo de senha"""
        self.password_field.clear()
        self.password_field.setPlaceholderText("Sua senha aparecerá aqui")

    def clean_old_records(self):
        """Limpa registros com mais de 60 dias da planilha"""
        try:
            excel_file = Path('senhas_database.xlsx')
            if not excel_file.exists():
                QMessageBox.warning(self, "Aviso", "Nenhuma planilha encontrada!")
                return

            # Tenta abrir a planilha
            try:
                wb = load_workbook(excel_file)
                ws = wb.active
            except PermissionError:
                QMessageBox.critical(self, "Erro", "A planilha está aberta em outro programa. Feche-a e tente novamente.")
                return

            # Confirma com o usuário
            resposta = QMessageBox.question(
                self,
                "Confirmar Limpeza",
                "Isso removerá permanentemente todos os registros com mais de 60 dias.\nDeseja continuar?",
                QMessageBox.Yes | QMessageBox.No
            )

            if resposta == QMessageBox.No:
                return

            # Data atual para comparação
            data_atual = datetime.now()
            registros_removidos = 0
            linhas_para_remover = []

            # Itera pelas linhas (de baixo para cima para não afetar os índices)
            for row in range(ws.max_row, 1, -1):  # Começa de max_row até 2 (ignora cabeçalho)
                data_cell = ws[f'D{row}'].value
                if data_cell:
                    try:
                        # Converte a data da planilha (formato dd/mm/yyyy)
                        data_registro = datetime.strptime(data_cell, "%d/%m/%Y")
                        dias_diferenca = (data_atual - data_registro).days

                        if dias_diferenca > 60:
                            linhas_para_remover.append(row)
                            registros_removidos += 1
                    except ValueError:
                        continue  # Ignora linhas com formato de data inválido

            # Remove as linhas marcadas
            for row in linhas_para_remover:
                ws.delete_rows(row)

            # Atualiza os IDs após a remoção
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
                row[0].value = idx  # Atualiza a coluna ID

            # Salva as alterações
            wb.save(excel_file)

            # Mostra mensagem de sucesso
            if registros_removidos > 0:
                QMessageBox.information(
                    self,
                    "Limpeza Concluída",
                    f"Foram removidos {registros_removidos} registros antigos com sucesso!"
                )
            else:
                QMessageBox.information(
                    self,
                    "Limpeza Concluída",
                    "Não foram encontrados registros com mais de 60 dias."
                )

            # Tenta abrir a planilha atualizada
            try:
                if sys.platform == 'win32':
                    os.startfile(excel_file)
                elif sys.platform == 'darwin':
                    os.system(f'open "{excel_file}"')
                else:
                    os.system(f'xdg-open "{excel_file}"')
            except:
                pass

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao limpar registros: {str(e)}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    
    # Configuração da fonte
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    
    window = PasswordGenerator()
    window.show()
    sys.exit(app.exec_())

















































